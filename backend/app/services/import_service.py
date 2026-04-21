import logging
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.anomaly_service import scan_anomalies
from app.services.aggregate_service import build_subtree_aggregates
from app.services.index_service import build_indexes
from app.services.parse_service import parse_rows_to_flat_nodes
from app.validators.workbook_validator import validate_workbook


logger = logging.getLogger(__name__)
ImportResult = dict[str, object]


def _sheet_headers(sheet: Worksheet) -> list[str]:
    return [str(cell.value).strip() for cell in sheet[1]]


def _sheet_rows(sheet: Worksheet, headers: list[str]) -> list[dict[str, object]]:
    return [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]


def import_dataset(file_obj: BytesIO) -> ImportResult:
    logger.info("开始加载 Excel 工作簿...")
    workbook = load_workbook(file_obj, data_only=True)
    logger.info(f"工作簿加载完成，工作表列表: {workbook.sheetnames}")

    logger.info("开始验证工作簿...")
    validation = validate_workbook(workbook)
    logger.info(f"验证完成，状态: {validation.status}")

    if validation.status == "failed":
        logger.warning(f"验证失败，错误数: {len(validation.errors)}")
        return {
            "status": "failed",
            "summary": validation.summary.model_dump(),
            "errors": [item.model_dump() for item in validation.errors],
        }

    sheet = workbook["子项明细"]
    headers = _sheet_headers(sheet)
    logger.info(f"表头读取完成，列数: {len(headers)}")
    logger.debug(f"表头: {headers}")

    rows = _sheet_rows(sheet, headers)
    logger.info(f"数据行读取完成，行数: {len(rows)}")

    logger.info("开始解析行数据...")
    flat_rows, parse_errors = parse_rows_to_flat_nodes(rows)
    logger.info(f"解析完成，有效行数: {len(flat_rows)}, 错误数: {len(parse_errors)}")

    if parse_errors:
        logger.warning(f"解析失败，错误详情: {parse_errors[:3]}...")  # 只打印前3个错误
        return {
            "status": "failed",
            "summary": {
                "fatal_count": len(parse_errors),
                "warning_count": 0,
            },
            "errors": parse_errors,
        }

    # 右侧分析区会频繁切换焦点节点，导入时预计算整棵子树统计可以保持交互稳定。
    logger.info("开始构建子树聚合数据...")
    subtree_aggregates = build_subtree_aggregates(flat_rows)
    logger.info(f"子树聚合完成，节点数: {len(subtree_aggregates)}")

    # 路径和 where-used 共享同一批 parent-child 关系，导入阶段一次构建可以避免后续重复遍历。
    logger.info("开始构建路径索引与 where-used 索引...")
    try:
        rows_with_indexes, indexes = build_indexes(flat_rows)
    except ValueError as exc:
        logger.warning(f"路径索引构建失败: {exc}")
        return {
            "status": "failed",
            "summary": {
                "fatal_count": 1,
                "warning_count": 0,
            },
            "errors": [
                {
                    "severity": "fatal",
                    "code": "INVALID_PARENT_ID",
                    "row_index": None,
                    "field": "parent_id",
                    "raw_value": "",
                    "message": str(exc),
                    "action": "请修正父子关系后重新导入",
                }
            ],
        }
    logger.info("路径索引构建完成")

    logger.info("开始扫描异常规则...")
    anomalies = scan_anomalies(rows_with_indexes)
    logger.info(f"异常扫描完成，异常数: {len(anomalies)}")

    return {
        "status": "success",
        "summary": {
            "total_rows": len(rows),
            "valid_rows": len(flat_rows),
            "warning_count": len(anomalies),
        },
        "rows": rows_with_indexes,
        "indexes": indexes,
        "subtree_aggregates": subtree_aggregates,
        # warnings 作为对外展示的异常摘要，和 anomalies 保持一致，避免详情页永远读到空列表。
        "warnings": anomalies,
        "anomalies": anomalies,
    }
